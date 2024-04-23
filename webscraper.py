import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Name of the company
name = input("Enter the name of the company: ")

# Stock symbol
sym = input("Enter the symbol for the company: ")

# URL of the webpage
url = "https://stockanalysis.com/stocks/" + sym + "/financials/ratios/"

# Send a GET request to the URL
response = requests.get(url)

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Find all <td> tags on the page
td_tags = soup.find_all(['td'])

# Filter out <td> tags containing the word "upgrade"
td_data = []
for tag in td_tags:
    text = tag.get_text(strip=True)
    if "upgrade" not in text.lower():
        td_data.append(text)

# Create a new Workbook
wb = Workbook()
ws = wb.active


# Find all <th> tags on the page and how many years there are recorded
th_tags = soup.find_all('th')
th_num = 0
for idx, th_tag in enumerate(th_tags, start=1):
    th_num +=1

# Write the "Name" column
ws.cell(row=1, column=1, value="Name")
# Write the company name for each year
for i in range(1, th_num + 1):
    ws.cell(row=i, column=1, value=name)

# Initialize counters
row_counter = 1
column_counter = 1

# Write the content of each <td> tag to the Excel sheet which represent the financial ratios
for idx, td_content in enumerate(td_data, start=1):
    # If the content represents a new category, create a new column
    if any(category in td_content for category in ["Market Capitalization", "Market Cap Growth", "Enterprise Value",
                                                    "PE Ratio", "PS Ratio", "PB Ratio", "P/FCF Ratio", "P/OCF Ratio",
                                                    "EV/Sales Ratio", "EV/EBITDA Ratio", "EV/EBIT Ratio",
                                                    "EV/FCF Ratio", "Debt / Equity Ratio", "Debt / EBITDA Ratio",
                                                    "Debt / FCF Ratio", "Quick Ratio", "Current Ratio",
                                                    "Asset Turnover", "Interest Coverage", "Return on Equity",
                                                    "Return on Assets", "Return on Capital", "Earnings Yield",
                                                    "FCF Yield", "Dividend Yield" ,"Payout Ratio", "Buyback Yield / Dilution", "Total Shareholder Return"]):
        row_counter = 1  # Reset the row counter for a new category
        column_counter += 1  # Move to the next column

    ws.cell(row=row_counter, column=column_counter, value=td_content)

    # Move to the next row
    row_counter += 1

# Writes the content of the <th> tags which represent the years
row_counter = 1
for idx, th_tag in enumerate(th_tags, start=1):
    th_content = th_tag.get_text(strip=True)
    ws.cell(row=row_counter, column=column_counter, value=th_content)
    
    # Move to the next row
    row_counter += 1

    
# Create the subfolder if it doesn't exist
subfolder = "Dataset"
if not os.path.exists(subfolder):
    os.makedirs(subfolder)

# Save the Excel file into the subfolder with the company name
file_path = os.path.join(subfolder, name + "_Ratios.xlsx")
wb.save(file_path)
