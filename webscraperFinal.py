
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Name of the company
name = input("Enter the name of the company: ")

# Stock symbol
sym = input("Enter the symbol for the company: ")
yearcount = int(input('How many years: ').strip() or "11")

# URL of the webpage
url = "https://stockanalysis.com/stocks/" + sym + "/financials/ratios/"

# Send a GET request to the URL
response = requests.get(url)

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Find all <td> tagsn the page
td_tags = soup.find_all('td') 

# Create a new Workbook
wb = Workbook()
ws = wb.active

# Write the "Name" column
ws.cell(row=1, column=1, value="Name")
for i in range(2,yearcount+2):
    ws.cell(row=i, column=1, value=name)  # Write the company name in the second row

# Initialize counters
row_counter = 1
column_counter = 2

# Write the content of each <td> tag to the Excel sheet which represent the financial ratios
for idx, td_tag in enumerate(td_tags, start=1):
    # Skip every 13th <td> tag
    if yearcount+1 == 12:
        if idx % 13 == 0: continue
    
    td_content = td_tag.get_text(strip=True)
    ws.cell(row=row_counter, column=column_counter, value=td_content)
    
    # Move to the next row
    row_counter += 1
    
    # If we've reached the 12th row, move to the next column and reset the row counter
    if row_counter > yearcount+1:
        row_counter = 1
        column_counter += 1

# Find all <th> tags on the page
th_tags = soup.find_all('th')


# Writes the content of the <th> tags which represent the years
row_counter = 1
for idx, th_tag in enumerate(th_tags, start=1):
    th_content = th_tag.get_text(strip=True)
    ws.cell(row=row_counter, column=column_counter, value=th_content)
    
    # Move to the next row
    row_counter += 1


# Create the subfolder if it doesn't exist
subfolder = "Dataset"
if not os.path.exists(subfolder):
    os.makedirs(subfolder)

# Save the Excel file into the subfolder with the company name
file_path = os.path.join(subfolder, name + "_Ratios.xlsx")
wb.save(file_path)