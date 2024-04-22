
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

sector = input("Please input the sector you would like to obtain: ")

# URL of the webpage
url = "https://stockanalysis.com/stocks/sector/" + sector + "/"

# Send a GET request to the URL
response = requests.get(url)

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Find all <td> tagsn the page
td_tags = soup.find_all('td') 

# Create a new Workbook
wb = Workbook()
ws = wb.active

# Initialize counters
row_counter = 2
column_counter = 1

# Write the content of each <td> tag to the Excel sheet which represent the financial ratios
for idx, td_tag in enumerate(td_tags, start=1):
    # Makes it stop after tthe top 200 companies of that sector
    if row_counter > 201:
        break
    # Only includes relevant important data, ignoring the rest
    if column_counter < 4:
        td_content = td_tag.get_text(strip=True)
        ws.cell(row=row_counter, column=column_counter, value=td_content)
    
    column_counter += 1
    # Move to the next row
    if column_counter == 8:
        column_counter  = 1
        row_counter +=1
    

# Find all <th> tags on the page
th_tags = soup.find_all('th')


# Writes the content of the <th> tags which represent the years
row_counter = 1
column_counter = 1
for idx, th_tag in enumerate(th_tags, start=1):
    if idx > 3:  # Check if the index is greater than 2
        break  # If so, exit the loop
    th_content = th_tag.get_text(strip=True)
    ws.cell(row=row_counter, column=column_counter, value=th_content)
    
    # Move to the next column
    column_counter += 1

    


# Create the subfolder if it doesn't exist
subfolder = "Dataset"
if not os.path.exists(subfolder):
    os.makedirs(subfolder)

# Save the Excel file into the subfolder with the company name
file_path = os.path.join(subfolder, sector + "SectorInformation.xlsx")
wb.save(file_path)