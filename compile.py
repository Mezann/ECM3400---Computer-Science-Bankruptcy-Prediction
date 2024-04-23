import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Define a mapping of column names to the types of data they represent
column_mapping = {
    "Name": ["Name"],
    "Market Capitalization": ["Market Capitalization", "Market Cap"],
    "Market Cap Growth": ["Market Cap Growth"],
    "Enterprise Value": ["Enterprise Value", "EV"],
    "PE Ratio": ["PE Ratio"],
    "PS Ratio": ["PS Ratio"],
    "PB Ratio": ["PB Ratio"],
    "P/FCF Ratio": ["P/FCF Ratio"],
    "P/OCF Ratio": ["P/OCF Ratio"],
    "EV/Sales Ratio": ["EV/Sales Ratio"],
    "EV/EBITDA Ratio": ["EV/EBITDA Ratio"],
    "EV/EBIT Ratio": ["EV/EBIT Ratio"],
    "EV/FCF Ratio": ["EV/FCF Ratio"],
    "Debt / Equity Ratio": ["Debt / Equity Ratio"],
    "Debt / EBITDA Ratio": ["Debt / EBITDA Ratio"],
    "Debt / FCF Ratio": ["Debt / FCF Ratio"],
    "Quick Ratio": ["Quick Ratio"],
    "Current Ratio": ["Current Ratio"],
    "Asset Turnover": ["Asset Turnover"],
    "Interest Coverage": ["Interest Coverage"],
    "Return on Equity (ROE)": ["Return on Equity (ROE)", "ROE"],
    "Return on Assets (ROA)": ["Return on Assets (ROA)", "ROA"],
    "Return on Capital (ROIC)": ["Return on Capital (ROIC)", "ROIC"],
    "Earnings Yield": ["Earnings Yield"],
    "FCF Yield": ["FCF Yield"],
    "Dividend Yield": ["Dividend Yield"],
    "Payout Ratio": ["Payout Ratio"],
    "Buyback Yield / Dilution": ["Buyback Yield / Dilution"],
    "Year": ["Year"]
}



def get_column_names(sheet):
    # Get the column names from the first row of the sheet
    column_names = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
        column_names.append(cell)
    return column_names

def compile_excel_files(input_folder, output_file):
    # Create a new workbook to compile all data
    compiled_wb = Workbook()
    compiled_ws = compiled_wb.active

    # Write headers based on column mapping
    headers = list(column_mapping.keys())
    compiled_ws.append(headers)

    # Create a new sheet for compiling data
    compiled_sheet = compiled_wb.create_sheet(title="Compiled Data")

    # Variable to track if any data is processed
    data_processed = False

    for file_name in os.listdir(input_folder):
        if file_name.endswith(".xlsx") and not file_name.startswith("~$"):
            file_path = os.path.join(input_folder, file_name)
            try:
                # Load the workbook
                workbook = load_workbook(file_path)
                for sheet in workbook.sheetnames:
                    # Get the column names from the sheet
                    sheet_columns = get_column_names(workbook[sheet])
                    # Copy data from each sheet to the compiled sheet
                    for row in workbook[sheet].iter_rows(min_row=2, values_only=True):
                        compiled_row = [''] * len(headers)
                        for i, cell_value in enumerate(row):
                            # Check the column name and place the value accordingly
                            for col_name, col_aliases in column_mapping.items():
                                if sheet_columns[i] in col_aliases:
                                    compiled_row[headers.index(col_name)] = cell_value
                        compiled_sheet.append(compiled_row)
                        data_processed = True
            except Exception as e:
                print(f"Error processing {file_name}: {e}")

    # Check if any data is processed before saving
    if data_processed:
        # Remove the default sheet created by Workbook()
        compiled_wb.remove(compiled_wb["Sheet"])
        # Save the compiled workbook to the output file
        compiled_wb.save(output_file)
    else:
        print("No data processed. Workbook not saved.")

# Example usage:
input_folder = "Dataset/communication-services"
output_file = "communication-servicesCompiled.xlsx"
compile_excel_files(input_folder, output_file)