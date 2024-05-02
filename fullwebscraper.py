import os
import requests
import openpyxl
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook

def sector_scraper(sector):
    ''' Obtains a list of all the companies in a chosen sector, picks the first 199 and puts them in an excel spreadsheet '''


    # URL of the webpage
    url = "https://stockanalysis.com/stocks/sector/" + sector + "/"

    # Send a GET request to the URL
    response = requests.get(url)

    # Parse the HTML content of the webpage
    soup = BeautifulSoup(response.text, 'html.parser')

    # Find all <td> tagsn the page
    td_tags = soup.find_all('td') 

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Initialize counters
    row_counter = 2
    column_counter = 1

    # Write the content of each <td> tag to the Excel sheet which represent the financial ratios
    for idx, td_tag in enumerate(td_tags, start=1):
        # Makes it stop after tthe top 200 companies of that sector
        if row_counter > 201:
            break
        # Only includes relevant important data, ignoring the rest
        if column_counter < 4:
            td_content = remove_symbols(td_tag.get_text(strip=True))
            ws.cell(row=row_counter, column=column_counter, value=td_content)
        
        column_counter += 1
        # Move to the next row
        if column_counter == 8:
            column_counter  = 1
            row_counter +=1
        

    # Find all <th> tags on the page
    th_tags = soup.find_all('th')


    # Writes the content of the <th> tags which represent the headings
    row_counter = 1
    column_counter = 1
    for idx, th_tag in enumerate(th_tags, start=1):
        if idx > 3:  # Check if the index is greater than 2
            break  # If so, exit the loop
        th_content = th_tag.get_text(strip=True)
        ws.cell(row=row_counter, column=column_counter, value=th_content)
        
        # Move to the next column
        column_counter += 1


    # Create the subfolder if it doesn't exist
    subfolder = "Dataset"
    if not os.path.exists(subfolder):
        os.makedirs(subfolder)

    # Save the Excel file into the subfolder with the company name
    file_path = os.path.join(subfolder, sector + "SectorInformation.xlsx")
    wb.save(file_path)

    return sector

def scrape_and_create_excel(sym, name, sector):
    '''Scrapes a specific company's financial ratios and puts it in an excel spreadsheet '''

    # URL of the webpage
    url = "https://stockanalysis.com/stocks/" + sym + "/financials/ratios/"

    # Send a GET request to the URL
    response = requests.get(url)

    # Parse the HTML content of the webpage
    soup = BeautifulSoup(response.text, 'html.parser')

    # Find all <td> tags on the page
    td_tags = soup.find_all(['td'])

    # Filter out <td> tags containing the word "upgrade"
    td_data = []
    for tag in td_tags:
        text = tag.get_text(strip=True)
        if "upgrade" not in text.lower():
            td_data.append(text)

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active


    # Find all <th> tags on the page and how many years there are recorded
    th_tags = soup.find_all('th')
    th_num = 0
    for idx, th_tag in enumerate(th_tags, start=1):
        th_num +=1

    # Write the "Name" column
    ws.cell(row=1, column=1, value="Name")
    # Write the company name for each year
    for i in range(2, th_num+1):
        ws.cell(row=i, column=1, value=name)

    # Initialize counters
    row_counter = 1
    column_counter = 1

    # Write the content of each <td> tag to the Excel sheet which represent the financial ratios
    for idx, td_content in enumerate(td_data, start=1):
        # If the content represents a new category, create a new column
        if any(category in td_content for category in ["Market Capitalization", "Market Cap Growth", "Enterprise Value",
                                                        "PE Ratio", "PS Ratio", "PB Ratio", "P/FCF Ratio", "P/OCF Ratio",
                                                        "EV/Sales Ratio", "EV/EBITDA Ratio", "EV/EBIT Ratio",
                                                        "EV/FCF Ratio", "Debt / Equity Ratio", "Debt / EBITDA Ratio",
                                                        "Debt / FCF Ratio", "Quick Ratio", "Current Ratio",
                                                        "Asset Turnover", "Interest Coverage", "Return on Equity",
                                                        "Return on Assets", "Return on Capital", "Earnings Yield",
                                                        "FCF Yield", "Dividend Yield" ,"Payout Ratio", "Buyback Yield / Dilution", "Total Shareholder Return"]):
            row_counter = 1  # Reset the row counter for a new category
            column_counter += 1  # Move to the next column

        ws.cell(row=row_counter, column=column_counter, value=td_content)

        # Move to the next row
        row_counter += 1

    # Writes the content of the <th> tags which represent the years
    row_counter = 1
    for idx, th_tag in enumerate(th_tags, start=1):
        th_content = th_tag.get_text(strip=True)
        ws.cell(row=row_counter, column=column_counter, value=th_content)
        
        # Move to the next row
        row_counter += 1

        
    # Create the subfolder for the sector if it doesn't exist
    sector_folder = os.path.join("Dataset", sector)
    if not os.path.exists(sector_folder):
        os.makedirs(sector_folder)

    # Save the Excel file into the subfolder with the company name
    file_path = os.path.join(sector_folder, name + " Ratios.xlsx")
    wb.save(file_path)


def sector_selection(sector):
    '''Compiles a spreadsheet of all the financial ratios from the top 199 companies in a specific sector'''

    # Define the file path relative to the current directory
    file_path = "Dataset/" + sector + "SectorInformation.xlsx"

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the active worksheet
    sheet = workbook.active

    ratio = 0
    # Iterate through rows, starting from the second row (index 1)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Access values in the second and third columns (index 1 and 2)
        sym = row[1]
        name = row[2]


        ratio += 1
        # Do something with the values
        scrape_and_create_excel(sym, name, sector)
        #print("Ratio " + str(ratio) + " made: " + name + " ratio")

def remove_symbols(name):
    # Define a regular expression pattern to match symbols excluding '.'
    pattern = r'[^\w\s.]'  # Matches any character that is not a letter, digit, whitespace, or '.'

    # Use the sub() function to replace all matches of the pattern with an empty string
    clean_text = re.sub(pattern, '', name)
    
    return clean_text


#sector_selection(sector_scraper("financials"))
#sector_selection(sector_scraper("technology"))
#sector_selection(sector_scraper("healthcare"))
#sector_selection(sector_scraper("industrials"))
#sector_selection(sector_scraper("consumer-discretionary"))
#sector_selection(sector_scraper("materials"))
#sector_selection(sector_scraper("real-estate"))
#sector_selection(sector_scraper("communication-services"))
#sector_selection(sector_scraper("energy"))
#sector_selection(sector_scraper("consumer-staples"))
#sector_selection(sector_scraper("utilities"))
#sector_scraper("communication-services")
sector_selection("technology")
#scrape_and_create_excel("SDCCQ","SmileDirectClub", "Bankrupcies")